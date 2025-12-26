#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
2026年工作日历生成器
根据模板生成3-12月份的工作日历，包含农历和节假日信息
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import date, datetime, timedelta
from copy import copy
import calendar
from lunarcalendar import Converter, Solar

# 2026年节假日安排
HOLIDAYS_2026 = {
    # 元旦: 1月1日-3日放假，1月4日上班
    date(2026, 1, 1): ("元旦", "休"),
    date(2026, 1, 2): ("元旦", "休"),
    date(2026, 1, 3): ("元旦", "休"),
    date(2026, 1, 4): ("元旦", "班"),
    
    # 春节: 2月15日-23日放假，2月14日、2月28日上班
    date(2026, 2, 14): ("春节", "班"),
    date(2026, 2, 15): ("春节", "休"),
    date(2026, 2, 16): ("春节", "休"),
    date(2026, 2, 17): ("春节", "休"),  # 除夕
    date(2026, 2, 18): ("春节", "休"),  # 初一
    date(2026, 2, 19): ("春节", "休"),  # 初二
    date(2026, 2, 20): ("春节", "休"),  # 初三
    date(2026, 2, 21): ("春节", "休"),
    date(2026, 2, 22): ("春节", "休"),
    date(2026, 2, 23): ("春节", "休"),
    date(2026, 2, 28): ("春节", "班"),
    
    # 清明节: 4月4日-6日放假
    date(2026, 4, 4): ("清明", "休"),
    date(2026, 4, 5): ("清明", "休"),
    date(2026, 4, 6): ("清明", "休"),
    
    # 劳动节: 5月1日-5日放假，5月9日上班
    date(2026, 5, 1): ("劳动节", "休"),
    date(2026, 5, 2): ("劳动节", "休"),
    date(2026, 5, 3): ("劳动节", "休"),
    date(2026, 5, 4): ("劳动节", "休"),
    date(2026, 5, 5): ("劳动节", "休"),
    date(2026, 5, 9): ("劳动节", "班"),
    
    # 端午节: 6月19日-21日放假
    date(2026, 6, 19): ("端午", "休"),
    date(2026, 6, 20): ("端午", "休"),
    date(2026, 6, 21): ("端午", "休"),
    
    # 中秋节: 9月25日-27日放假
    date(2026, 9, 25): ("中秋", "休"),
    date(2026, 9, 26): ("中秋", "休"),
    date(2026, 9, 27): ("中秋", "休"),
    
    # 国庆节: 10月1日-7日放假，9月20日、10月10日上班
    date(2026, 9, 20): ("国庆", "班"),
    date(2026, 10, 1): ("国庆", "休"),
    date(2026, 10, 2): ("国庆", "休"),
    date(2026, 10, 3): ("国庆", "休"),
    date(2026, 10, 4): ("国庆", "休"),
    date(2026, 10, 5): ("国庆", "休"),
    date(2026, 10, 6): ("国庆", "休"),
    date(2026, 10, 7): ("国庆", "休"),
    date(2026, 10, 10): ("国庆", "班"),
}

# 农历数字转中文
LUNAR_DAYS = ['初一', '初二', '初三', '初四', '初五', '初六', '初七', '初八', '初九', '初十',
              '十一', '十二', '十三', '十四', '十五', '十六', '十七', '十八', '十九', '二十',
              '廿一', '廿二', '廿三', '廿四', '廿五', '廿六', '廿七', '廿八', '廿九', '三十']

LUNAR_MONTHS = ['正月', '二月', '三月', '四月', '五月', '六月', 
                '七月', '八月', '九月', '十月', '冬月', '腊月']

def get_lunar_date(solar_date):
    """获取农历日期"""
    solar = Solar(solar_date.year, solar_date.month, solar_date.day)
    lunar = Converter.Solar2Lunar(solar)
    
    day_str = LUNAR_DAYS[lunar.day - 1]
    # 如果是初一，显示月份
    if lunar.day == 1:
        month_str = LUNAR_MONTHS[lunar.month - 1]
        if lunar.isleap:
            month_str = "闰" + month_str
        return month_str
    return day_str

def get_holiday_info(d):
    """获取节假日信息"""
    if d in HOLIDAYS_2026:
        name, status = HOLIDAYS_2026[d]
        return f"{name}（{status}）"
    return ""

def copy_cell_style(source_cell, target_cell):
    """复制单元格样式"""
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)

def get_weeks_for_month(year, month):
    """
    获取月份的周数据，每周从周一开始
    返回: [(week_start_date, week_end_date), ...]
    包含该月第一天所在的周到最后一天所在的周
    """
    # 获取该月第一天
    first_day = date(year, month, 1)
    # 获取该月最后一天
    if month == 12:
        last_day = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        last_day = date(year, month + 1, 1) - timedelta(days=1)
    
    # 找到第一天所在周的周一
    days_since_monday = first_day.weekday()  # 0 = Monday
    week_start = first_day - timedelta(days=days_since_monday)
    
    weeks = []
    while week_start <= last_day:
        week_end = week_start + timedelta(days=6)
        weeks.append((week_start, week_end))
        week_start = week_end + timedelta(days=1)
    
    return weeks

def create_month_sheet(wb, template_sheet, year, month):
    """根据模板创建月份工作表"""
    sheet_name = f"2026-{month:02d}"
    
    # 复制模板sheet
    new_sheet = wb.copy_worksheet(template_sheet)
    new_sheet.title = sheet_name
    
    # 获取该月的所有周
    weeks = get_weeks_for_month(year, month)
    
    # 更新月份标题 (B1单元格)
    new_sheet['B1'] = f"2026 年 {month} 月"
    
    # 行号配置 - 根据模板结构
    # Row 2: 星期标题头 (不需要修改)
    # Row 3, 5, 7, 9, 11, 13: 日期行 (每周一行)
    # Row 4, 6, 8, 10, 12, 14: 信息行 (农历/节假日)
    
    date_rows = [3, 5, 7, 9, 11, 13]  # 日期行
    info_rows = [4, 6, 8, 10, 12, 14]  # 信息行
    
    # 清空现有数据（从第3行开始到第16行，确保没有残留数据）
    for row in range(3, 17):
        for col in range(1, 9):  # A到H列
            new_sheet.cell(row=row, column=col, value='')
    
    # 填充每周数据
    for week_idx, (week_start, week_end) in enumerate(weeks):
        if week_idx >= len(date_rows):
            break  # 最多6周
        
        date_row = date_rows[week_idx]
        info_row = info_rows[week_idx]
        
        # 填充周一到周日 (B-H列，对应列2-8)
        for day_offset in range(7):
            current_date = week_start + timedelta(days=day_offset)
            col = day_offset + 2  # B=2, C=3, ..., H=8
            
            # 写入日期 (datetime格式)
            cell_date = new_sheet.cell(row=date_row, column=col)
            cell_date.value = datetime(current_date.year, current_date.month, current_date.day)
            
            # 获取农历和节假日信息
            lunar_str = get_lunar_date(current_date)
            holiday_info = get_holiday_info(current_date)
            
            # 写入信息行
            cell_info = new_sheet.cell(row=info_row, column=col)
            if holiday_info:
                cell_info.value = holiday_info
            else:
                # 如果不是当月的日期，可以显示农历
                cell_info.value = lunar_str
        
        # 在A列写入该周的农历信息 (可选：显示周的农历范围或留空)
        # 根据参考图片，A列应该显示农历日期
        first_day_of_week = week_start
        lunar_week_info = get_lunar_date(first_day_of_week)
        new_sheet.cell(row=date_row, column=1, value=lunar_week_info)
    
    return new_sheet

def main():
    # 读取模板
    template_path = '/Users/buoge/Desktop/github/aicoding-bootcamp/w4/calendar-generate/2026年日历模板.xlsx'
    output_path = '/Users/buoge/Desktop/github/aicoding-bootcamp/w4/calendar-generate/2026年工作日历.xlsx'
    
    wb = openpyxl.load_workbook(template_path)
    
    # 获取模板sheet (使用2026-01作为模板)
    template_sheet = wb['2026-01']
    
    print("开始生成2026年3-12月工作日历...")
    print(f"模板sheets: {wb.sheetnames}")
    
    # 需要保留的sheets
    sheets_to_keep = ['2026-01', '2026-02']
    
    # 生成3-12月的日历
    for month in range(3, 13):
        print(f"生成 {month} 月...")
        create_month_sheet(wb, template_sheet, 2026, month)
    
    # 删除不需要的sheets（保留2026-01, 2026-02和新生成的月份）
    sheets_to_remove = []
    for sheet_name in wb.sheetnames:
        if sheet_name not in sheets_to_keep and not sheet_name.startswith('2026-'):
            sheets_to_remove.append(sheet_name)
        # 检查是否是有效的2026月份sheet
        if sheet_name.startswith('2026-'):
            try:
                month_part = sheet_name.split('-')[1]
                month_num = int(month_part)
                if month_num < 1 or month_num > 12:
                    sheets_to_remove.append(sheet_name)
            except:
                pass
    
    for sheet_name in sheets_to_remove:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
    
    # 重新排序sheets
    # 获取所有2026月份sheets并排序
    month_sheets = []
    for sheet_name in wb.sheetnames:
        if sheet_name.startswith('2026-'):
            try:
                month_num = int(sheet_name.split('-')[1])
                month_sheets.append((month_num, sheet_name))
            except:
                pass
    month_sheets.sort(key=lambda x: x[0])
    
    # 保存文件
    wb.save(output_path)
    print(f"\n日历已保存到: {output_path}")
    print(f"共生成 {len(month_sheets)} 个月份的日历")

if __name__ == "__main__":
    main()
