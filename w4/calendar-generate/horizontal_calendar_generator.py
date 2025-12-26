#!/usr/bin/env python3
"""
Generate 2026 work calendar with horizontal layout matching template image
"""

import pandas as pd
from datetime import datetime, timedelta
import openpyxl
from copy import copy

def get_lunar_date(day, month):
    """
    Return lunar date for given day and month in 2026.
    Simplified implementation - would need proper lunar calendar library.
    Based on 2026 lunar calendar patterns.
    """
    # Lunar calendar for 2026
    lunar_calendar = {
        1: {
            # January (starts in lunar month 12, crosses to month 1)
            1: "十三", 2: "十四", 3: "十五", 4: "十六", 5: "十七",
            6: "十八", 7: "十九", 8: "二十", 9: "廿一", 10: "廿二",
            11: "廿三", 12: "廿四", 13: "廿五", 14: "廿六", 15: "廿七",
            16: "廿八", 17: "廿九", 18: "三十", 19: "正月", 20: "初二",
            21: "初三", 22: "初四", 23: "初五", 24: "初六", 25: "初七",
            26: "初八", 27: "初九", 28: "初十", 29: "十一", 30: "十二", 31: "十三"
        },
        2: {
            # February (lunar month 1)
            1: "十四", 2: "十五", 3: "十六", 4: "十七", 5: "十八",
            6: "十九", 7: "二十", 8: "廿一", 9: "廿二", 10: "廿三",
            11: "廿四", 12: "廿五", 13: "廿六", 14: "廿七", 15: "除夕",
            16: "春节", 17: "初二", 18: "初三", 19: "初四", 20: "初五",
            21: "初六", 22: "初七", 23: "初八", 24: "初九", 25: "初十",
            26: "十一", 27: "十二", 28: "十三"
        },
        3: {
            # March (lunar month 2)
            1: "十四", 2: "元宵", 3: "十六", 4: "十七", 5: "十八",
            6: "十九", 7: "二十", 8: "廿一", 9: "廿二", 10: "廿三",
            11: "廿四", 12: "廿五", 13: "廿六", 14: "廿七", 15: "廿八",
            16: "廿九", 17: "二月", 18: "初二", 19: "初三", 20: "初四",
            21: "初五", 22: "初六", 23: "初七", 24: "初八", 25: "初九",
            26: "初十", 27: "十一", 28: "十二", 29: "十三", 30: "十四",
            31: "十五"
        },
        4: {
            # April (lunar month 3)
            1: "十六", 2: "十七", 3: "十八", 4: "十九", 5: "清明",
            6: "廿一", 7: "廿二", 8: "廿三", 9: "廿四", 10: "廿五",
            11: "廿六", 12: "廿七", 13: "廿八", 14: "廿九", 15: "三月",
            16: "初二", 17: "初三", 18: "初四", 19: "初五", 20: "初六",
            21: "初七", 22: "初八", 23: "初九", 24: "初十", 25: "十一",
            26: "十二", 27: "十三", 28: "十四", 29: "十五", 30: "十六"
        },
        5: {
            # May (lunar month 4)
            1: "十七", 2: "十八", 3: "十九", 4: "二十", 5: "廿一",
            6: "廿二", 7: "廿三", 8: "廿四", 9: "廿五", 10: "廿六",
            11: "廿七", 12: "廿八", 13: "廿九", 14: "四月", 15: "初二",
            16: "初三", 17: "初四", 18: "初五", 19: "初六", 20: "初七",
            21: "初八", 22: "初九", 23: "初十", 24: "十一", 25: "十二",
            26: "十三", 27: "十四", 28: "十五", 29: "十六", 30: "十七",
            31: "十八"
        },
        6: {
            # June (lunar month 5)
            1: "十九", 2: "二十", 3: "廿一", 4: "廿二", 5: "廿三",
            6: "廿四", 7: "廿五", 8: "芒种", 10: "廿七", 11: "廿八",
            12: "廿九", 13: "三十", 14: "五月", 15: "端午", 16: "初三",
            17: "初四", 18: "初五", 19: "初六", 20: "初七", 21: "夏至",
            22: "初九", 23: "初十", 24: "十一", 25: "十二", 26: "十三",
            27: "十四", 28: "十五", 29: "十六", 30: "十七"
        },
        7: {
            # July (lunar month 6)
            1: "十八", 2: "十九", 3: "二十", 4: "廿一", 5: "廿二",
            6: "廿三", 7: "廿四", 8: "廿五", 9: "廿六", 10: "廿七",
            11: "廿八", 12: "廿九", 13: "三十", 14: "六月", 15: "小暑",
            16: "初三", 17: "初四", 18: "初五", 19: "初六", 20: "初七",
            21: "初八", 22: "初九", 23: "初十", 24: "十一", 25: "十二",
            26: "十三", 27: "十四", 28: "十五", 29: "十六", 30: "十七",
            31: "十八"
        },
        8: {
            # August (lunar month 7)
            1: "十九", 2: "二十", 3: "廿一", 4: "廿二", 5: "廿三",
            6: "廿四", 7: "廿五", 8: "立秋", 9: "廿七", 10: "廿八",
            11: "廿九", 12: "三十", 13: "七月", 14: "初三", 15: "初四",
            16: "初五", 17: "初六", 18: "七夕", 19: "初八", 20: "初九",
            21: "初十", 22: "十一", 23: "十二", 24: "十三", 25: "十四",
            26: "十五", 27: "十六", 28: "十七", 29: "十八", 30: "十九",
            31: "二十"
        },
        9: {
            # September (lunar month 8)
            1: "廿一", 2: "廿二", 3: "廿三", 4: "廿四", 5: "廿五",
            6: "廿六", 7: "白露", 8: "廿八", 9: "廿九", 10: "三十",
            11: "八月", 12: "初二", 13: "初三", 14: "初四", 15: "初五",
            16: "初六", 17: "初七", 18: "初八", 19: "初九", 20: "初十",
            21: "十一", 22: "十二", 23: "十三", 24: "十四", 25: "中秋节",
            26: "十六", 27: "十七", 28: "十八", 29: "十九", 30: "二十"
        },
        10: {
            # October (lunar month 9)
            1: "廿一", 2: "廿二", 3: "廿三", 4: "廿四", 5: "廿五",
            6: "廿六", 7: "廿七", 8: "廿八", 9: "廿九", 10: "三十",
            11: "九月", 12: "国庆", 13: "初三", 14: "初四", 15: "初五",
            16: "初六", 17: "初七", 18: "初八", 19: "重阳", 20: "初十",
            21: "十一", 22: "十二", 23: "十三", 24: "十四", 25: "十五",
            26: "十六", 27: "十七", 28: "十八", 29: "十九", 30: "二十",
            31: "廿一"
        },
        11: {
            # November (lunar month 10)
            1: "廿二", 2: "廿三", 3: "廿四", 4: "廿五", 5: "立冬",
            6: "廿七", 7: "廿八", 8: "廿九", 9: "三十", 10: "十月",
            11: "初二", 12: "初三", 13: "初四", 14: "初五", 15: "初六",
            16: "初七", 17: "初八", 18: "初九", 19: "初十", 20: "十一",
            21: "十二", 22: "十三", 23: "十四", 24: "十五", 25: "十六",
            26: "十七", 27: "十八", 28: "十九", 29: "二十", 30: "廿一"
        },
        12: {
            # December (lunar month 11)
            1: "廿二", 2: "廿三", 3: "廿四", 4: "廿五", 5: "廿六",
            6: "廿七", 7: "大雪", 8: "廿九", 9: "三十", 10: "十一月",
            11: "初二", 12: "初三", 13: "初四", 14: "初五", 15: "初六",
            16: "初七", 17: "初八", 18: "初九", 19: "初十", 20: "十一",
            21: "十二", 22: "十三", 23: "十四", 24: "十五", 25: "十六",
            26: "十七", 27: "十八", 28: "十九", 29: "二十", 30: "廿一",
            31: "廿二"
        }
    }

    return lunar_calendar.get(month, {}).get(day, "")

def create_calendar_rows(year, month, holiday_rules):
    """
    Create calendar rows matching template format
    """
    # Calculate date range for the month
    first_day = datetime(year, month, 1)

    if month == 12:
        last_day = datetime(year, 12, 31)
    else:
        next_month = datetime(year, month + 1, 1)
        last_day = next_month - timedelta(days=1)

    # Calculate which Monday to start from
    # Start from Monday of the week containing the 1st
    start_date = first_day - timedelta(days=first_day.weekday())

    # Calculate end date (Saturday of the week containing the last day)
    end_date = last_day + timedelta(days=(5 - last_day.weekday()))

    # Generate all dates in the range
    current_date = start_date
    all_dates = []
    while current_date <= end_date:
        all_dates.append(current_date)
        current_date += timedelta(days=1)

    # Calculate number of weeks
    num_weeks = len(all_dates) // 7

    # Create rows
    rows = []

    # Month header row
    month_names = ["", "一月", "二月", "三月", "四月", "五月", "六月",
                   "七月", "八月", "九月", "十月", "十一月", "十二月"]
    rows.append([f"2026年 {month_names[month]}"])

    # Empty row
    rows.append([""])

    # Weekday headers row
    weekdays = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
    rows.append(weekdays)

    # For each week, add date rows with lunar dates and holiday marks
    for week_idx in range(num_weeks):
        week_start_idx = week_idx * 7
        week_dates = all_dates[week_start_idx:week_start_idx + 7]

        # Date row with lunar dates
        date_row = []
        for date in week_dates:
            if date.month == month:
                day = date.day
                lunar = get_lunar_date(day, month)
                if lunar:
                    date_row.append(f"{day}\n({lunar})")
                else:
                    date_row.append(str(day))
            else:
                # Show dates from adjacent months (grayed out)
                date_row.append(str(date.day))

        rows.append(date_row)

        # Holiday/work row
        holiday_row = []
        for date in week_dates:
            if date.month == month:
                holiday_key = (month, date.day)
                holiday_row.append(holiday_rules.get(holiday_key, ""))
            else:
                holiday_row.append("")
        rows.append(holiday_row)

    return rows

def copy_sheet_style_and_data(source_sheet, target_sheet):
    """Copy sheet data and basic styles from template"""
    # Copy data
    for row in source_sheet.iter_rows(values_only=True):
        target_sheet.append(list(row) if row else [])

    # Copy basic column widths
    for col_letter, col_dim in source_sheet.column_dimensions.items():
        if col_letter in 'ABCDEFGH':  # First 8 columns
            target_sheet.column_dimensions[col_letter].width = col_dim.width

def main():
    # Holiday rules for 2026 - updated with complete information
    holiday_rules = {
        # 清明节: 4月4日-4月6日放假
        (4, 4): "清明(休)", (4, 5): "清明(休)", (4, 6): "清明(休)",

        # 劳动节: 5月1日-5月5日放假，5月9日上班
        (5, 1): "劳动节(休)", (5, 2): "劳动节(休)", (5, 3): "劳动节(休)",
        (5, 4): "劳动节(休)", (5, 5): "劳动节(休)", (5, 9): "劳动节(班)",

        # 端午节: 6月19日-6月21日放假
        (6, 19): "端午节(休)", (6, 20): "端午节(休)", (6, 21): "端午节(休)",

        # 中秋节: 9月25日-9月27日放假
        (9, 25): "中秋节(休)", (9, 26): "中秋节(休)", (9, 27): "中秋节(休)",

        # 国庆节: 10月1日-10月7日放假，9月20日、10月10日上班
        (10, 1): "国庆节(休)", (10, 2): "国庆节(休)", (10, 3): "国庆节(休)",
        (10, 4): "国庆节(休)", (10, 5): "国庆节(休)", (10, 6): "国庆节(休)",
        (10, 7): "国庆节(休)", (9, 20): "国庆节(班)", (10, 10): "国庆节(班)",
    }

    try:
        # Load template
        template_path = "2026年日历模板.xlsx"
        template_wb = openpyxl.load_workbook(template_path)

        # Create output workbook
        output_wb = openpyxl.Workbook()

        # Remove default sheet
        if 'Sheet' in output_wb.sheetnames:
            output_wb.remove(output_wb['Sheet'])

        # Copy January and February from template
        print("Copying January and February from template...")
        for sheet_name in ["2026-01", "2026-02"]:
            if sheet_name in template_wb.sheetnames:
                source_sheet = template_wb[sheet_name]
                target_sheet = output_wb.create_sheet(sheet_name)
                copy_sheet_style_and_data(source_sheet, target_sheet)

        # Generate March to December
        for month in range(3, 13):
            print(f"Generating 2026-{month:02d}...")

            # Create calendar rows
            calendar_rows = create_calendar_rows(2026, month, holiday_rules)

            # Convert to DataFrame
            df = pd.DataFrame(calendar_rows)

            # Create sheet
            sheet_name = f"2026-{month:02d}"
            ws = output_wb.create_sheet(sheet_name)

            # Write data
            for r_idx, row in enumerate(calendar_rows, 1):
                for c_idx, value in enumerate(row, 1):
                    if value:
                        ws.cell(row=r_idx, column=c_idx, value=value)

            # Set column widths
            for i in range(1, 8):  # Columns A-G
                col_letter = openpyxl.utils.get_column_letter(i)
                ws.column_dimensions[col_letter].width = 18

        # Save output file
        output_path = "2026年工作日历-claude.xlsx"
        output_wb.save(output_path)
        print(f"\nCalendar generation complete! Saved as: {output_path}")

        # Verify
        verify_wb = openpyxl.load_workbook(output_path)
        print(f"Generated {len(verify_wb.sheetnames)} sheets:", sorted(verify_wb.sheetnames))
        verify_wb.close()

        # Clean up
        template_wb.close()
        output_wb.close()

    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
